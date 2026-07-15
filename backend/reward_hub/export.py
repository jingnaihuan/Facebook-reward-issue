# -*- coding: utf-8 -*-
"""导出发奖名单为多 sheet xlsx。"""
import openpyxl

PLAYER_COLS = ["留言顺序", "留言时间", "玩家ID", "语言", "别墅等级",
               "角色名称", "角色等级", "角色创建时间", "服务器",
               "历史充值总额", "最后登录时间"]
# 玩家 dict 字段 → 中文列
_FIELD = ["order", "time", "player_id", "lang", "villa", "role_name",
          "role_level", "role_created", "server", "total_recharge", "last_login"]


def _write_player_sheet(ws, rows, with_content=False):
    # 关键词奖多加一列「留言内容」，人工好核对『这些人确实答对了』。
    cols = PLAYER_COLS + (["留言内容"] if with_content else [])
    ws.append(cols)
    for r in rows:
        vals = [r.get(f, "") for f in _FIELD]
        if with_content:
            vals.append(r.get("content", ""))
        ws.append(vals)


def _write_invalid_sheet(ws, rows):
    header = ["玩家ID", "留言内容", "原因"]
    ws.append(header)
    for r in rows:
        ws.append([r.get("player_id", ""), r.get("content", ""),
                   r.get("reject_reason", "")])


def export_reward_workbook(path, awards, participation, invalid,
                           allow_winner_participation=False,
                           keyword_award_names=None):
    """awards: {奖项名: [玩家]}；participation: [玩家]；invalid: [无效记录]。
    allow_winner_participation：参与奖是否含抽选中奖者（仅用于 sheet 标题标注）。
    keyword_award_names：关键词奖名集合，这些 sheet 追加「留言内容」列供复核。"""
    kw_names = set(keyword_award_names or ())
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, winners in awards.items():
        _write_player_sheet(wb.create_sheet(title=name[:31]), winners,
                            with_content=name in kw_names)
    # 无任何抽选奖项 = 普惠奖（全员）：这些人是发奖对象，工作簿名标清；
    # 有抽选奖项时，这一档是落选者的「参与奖」；若允许中奖者重复领，则标注「含中奖者」。
    if not awards:
        part_title = "普惠奖（全员）"
    elif allow_winner_participation:
        part_title = "参与奖（含中奖者）"
    else:
        part_title = "参与奖"
    _write_player_sheet(wb.create_sheet(title=part_title), participation)
    _write_invalid_sheet(wb.create_sheet(title="无效"), invalid)
    wb.save(path)
    return path
