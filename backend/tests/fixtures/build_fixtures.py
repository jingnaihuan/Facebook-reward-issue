# -*- coding: utf-8 -*-
"""从「寻找伤心松鼠」xlsm 抽取回归 fixture（一次性运行，产物提交仓库）。"""
import json, os
import openpyxl

XLSM = os.path.expanduser(
    "~/Library/Containers/com.xunmeng.knock/5azlYjzeJT0A/files/寻找伤心松鼠社群互动发奖.xlsm")
HERE = os.path.dirname(os.path.abspath(__file__))

wb = openpyxl.load_workbook(XLSM, read_only=True, data_only=True)

# 原始留言 → raw_comments.json（序号/留言内容/按赞数/留言时间）
raw = []
ws = wb["原始留言"]
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    seq, name, content, likes, ctime, _ = row
    if seq is None:
        continue
    raw.append({"order": int(seq), "content": content or "",
                "likes": int(likes) if isinstance(likes, (int, float)) else 0,
                "replies": 0, "time": str(ctime) if ctime else ""})
json.dump(raw, open(os.path.join(HERE, "raw_comments.json"), "w"),
          ensure_ascii=False, indent=2)

# ID提取 sheet → expected_players.json（玩家ID → 玩家信息）打桩 Eastblue
players = {}
ws = wb["ID提取"]
for row in ws.iter_rows(min_row=2, values_only=True):
    pid = row[5]
    if not pid:
        continue
    players[str(pid)] = {
        "player_id": str(pid), "lang": row[6], "villa": row[7],
        "role_name": row[8], "role_level": row[9], "role_created": str(row[10]),
        "server": row[11], "total_recharge": row[12], "last_login": str(row[13])}
json.dump(players, open(os.path.join(HERE, "expected_players.json"), "w"),
          ensure_ascii=False, indent=2)

# 先锋奖/参与奖玩家ID集合（期望结果）
for sheet, fn in [("先锋奖", "expected_vanguard.json"),
                  ("参与奖", "expected_participation.json")]:
    ids = []
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[5]:
            ids.append(str(row[5]))
    json.dump(ids, open(os.path.join(HERE, fn), "w"), ensure_ascii=False, indent=2)

print("fixtures 生成完成")
