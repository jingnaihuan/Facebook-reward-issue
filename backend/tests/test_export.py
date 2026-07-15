import openpyxl
from reward_hub.export import export_reward_workbook

PLAYER_COLS = ["留言顺序", "留言时间", "玩家ID", "语言", "别墅等级",
               "角色名称", "角色等级", "角色创建时间", "服务器",
               "历史充值总额", "最后登录时间"]


def _p(pid, order, lang="en"):
    return {"order": order, "time": "2026-06-26 08:00:00", "player_id": pid,
            "lang": lang, "villa": "37", "role_name": "Hero", "role_level": "55",
            "role_created": "2019-05-02", "server": "S177",
            "total_recharge": "100", "last_login": "2026-07-03"}


def test_creates_sheets(tmp_path):
    out = tmp_path / "out.xlsx"
    awards = {"先锋奖": [_p("1000000001", 1)]}
    participation = [_p("1000000002", 2)]
    invalid = [{"player_id": "", "content": "no id", "reject_reason": "无有效ID"}]
    export_reward_workbook(str(out), awards, participation, invalid)

    wb = openpyxl.load_workbook(str(out))
    assert wb.sheetnames == ["先锋奖", "参与奖", "无效"]


def test_award_sheet_header_and_row(tmp_path):
    out = tmp_path / "out.xlsx"
    export_reward_workbook(str(out), {"先锋奖": [_p("1000000001", 1)]}, [], [])
    wb = openpyxl.load_workbook(str(out))
    ws = wb["先锋奖"]
    assert [c.value for c in ws[1]] == PLAYER_COLS
    assert ws.cell(row=2, column=3).value == "1000000001"  # 玩家ID 列


def test_universal_names_participation_sheet(tmp_path):
    """普惠奖（awards 为空）时，全员这一档工作簿名为「普惠奖（全员）」，且不含任何奖项 sheet。"""
    out = tmp_path / "out.xlsx"
    export_reward_workbook(str(out), {}, [_p("1000000001", 1), _p("1000000002", 2)], [])
    wb = openpyxl.load_workbook(str(out))
    assert wb.sheetnames == ["普惠奖（全员）", "无效"]


def test_normal_mode_keeps_participation_name(tmp_path):
    """有抽选奖项时，落选档仍叫「参与奖」（不改动既有正常发奖行为）。"""
    out = tmp_path / "out.xlsx"
    export_reward_workbook(str(out), {"先锋奖": [_p("1000000001", 1)]}, [_p("1000000002", 2)], [])
    wb = openpyxl.load_workbook(str(out))
    assert wb.sheetnames == ["先锋奖", "参与奖", "无效"]


def test_participation_sheet_labeled_when_winners_included(tmp_path):
    """开「中奖者可重复领参与奖」时，参与奖 sheet 标题标注「含中奖者」，
    避免拿到 xlsx 的人误以为参与奖里都是未中奖者。"""
    out = tmp_path / "out.xlsx"
    export_reward_workbook(str(out), {"先锋奖": [_p("1000000001", 1)]},
                           [_p("1000000001", 1), _p("1000000002", 2)], [],
                           allow_winner_participation=True)
    wb = openpyxl.load_workbook(str(out))
    assert wb.sheetnames == ["先锋奖", "参与奖（含中奖者）", "无效"]


def test_invalid_sheet_has_reason_column(tmp_path):
    out = tmp_path / "out.xlsx"
    invalid = [{"player_id": "", "content": "hi", "reject_reason": "无有效ID"}]
    export_reward_workbook(str(out), {}, [], invalid)
    wb = openpyxl.load_workbook(str(out))
    ws = wb["无效"]
    header = [c.value for c in ws[1]]
    assert "原因" in header
    assert ws.cell(row=2, column=header.index("原因") + 1).value == "无有效ID"


def test_keyword_award_sheet_has_content_column(tmp_path):
    """关键词奖 sheet 追加「留言内容」列，供人工复核『确实答对了』。"""
    out = tmp_path / "out.xlsx"
    winner = {**_p("1000000001", 1), "content": "我猜红色 1000000001"}
    export_reward_workbook(str(out), {"答题奖": [winner]}, [], [],
                           keyword_award_names={"答题奖"})
    wb = openpyxl.load_workbook(str(out))
    ws = wb["答题奖"]
    header = [c.value for c in ws[1]]
    assert header == PLAYER_COLS + ["留言内容"]
    assert ws.cell(row=2, column=len(header)).value == "我猜红色 1000000001"


def test_non_keyword_award_sheet_has_no_content_column(tmp_path):
    """普通奖 sheet 不加留言列（不改动既有导出列）。"""
    out = tmp_path / "out.xlsx"
    export_reward_workbook(str(out), {"盖楼奖": [_p("1000000001", 1)]}, [], [],
                           keyword_award_names={"答题奖"})   # 盖楼奖不在关键词集合里
    wb = openpyxl.load_workbook(str(out))
    assert [c.value for c in wb["盖楼奖"][1]] == PLAYER_COLS
